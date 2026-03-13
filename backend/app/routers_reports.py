from typing import Optional
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .auth import get_db, get_current_user
from .models import Equipment, User
from .permissions import require_manager_or_above

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


router = APIRouter(prefix="/reports", tags=["reports"])


# =========================================================
# EQUIPMENT EXCEL GENERATOR
# =========================================================
def generate_equipment_excel(
    db: Session,
    department_id: Optional[str] = None,
    status: Optional[str] = None,
):

    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel generation not available. Install openpyxl.",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"

    # Header styling
    header_fill = PatternFill(start_color="4472C4",
                              end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Asset Tag",
        "Equipment Name",
        "Manufacturer",
        "Model",
        "Serial Number",
        "Status",
        "Department",
        "Repair Count",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    query = db.query(Equipment)

    if department_id:
        query = query.filter(Equipment.department_id == department_id)

    if status:
        query = query.filter(Equipment.status == status)

    equipment_list = query.order_by(Equipment.equipment_name).all()

    for row, eq in enumerate(equipment_list, 2):

        department_name = eq.department.name if eq.department else "N/A"

        ws.cell(row=row, column=1, value=eq.asset_tag)
        ws.cell(row=row, column=2, value=eq.equipment_name)
        ws.cell(row=row, column=3, value=eq.manufacturer or "N/A")
        ws.cell(row=row, column=4, value=eq.model or "N/A")
        ws.cell(row=row, column=5, value=eq.serial_number or "N/A")
        ws.cell(row=row, column=6, value=str(eq.status))
        ws.cell(row=row, column=7, value=department_name)
        ws.cell(row=row, column=8, value=eq.repair_count)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# =========================================================
# DOWNLOAD EQUIPMENT REPORT
# =========================================================
@router.get("/equipment/excel")
def download_equipment_report(
    department_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    require_manager_or_above(current_user)

    output = generate_equipment_excel(db, department_id, status)

    filename = f"equipment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
